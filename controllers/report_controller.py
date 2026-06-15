"""ReportController - xuất dữ liệu thống kê sang Excel hoặc CSV."""

import csv
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None

from controllers.database_connection import DatabaseConnection


class ReportController:
    """Xử lý truy vấn, định dạng và xuất báo cáo dữ liệu."""

    def __init__(self):
        self.db = DatabaseConnection()
        self.export_dir = "reports"
        self._ensure_export_dir_exists()

    def _ensure_export_dir_exists(self):
        """Tạo thư mục lưu trữ báo cáo nếu chưa tồn tại."""
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)

    def export_invoice_to_csv(self, record_id):
        """Truy xuất chi tiết một đơn mượn và xuất thành hóa đơn (Excel/CSV)."""
        query = """
            SELECT u.full_name, br.borrow_date, br.expected_return_date, 
                   d.device_name, bd.quantity, bd.returned
            FROM borrow_records br
            JOIN users u ON br.user_id = u.user_id
            JOIN borrow_details bd ON br.record_id = bd.record_id
            JOIN devices d ON bd.device_id = d.device_id
            WHERE br.record_id = ?
        """
        try:
            results = self.db.execute_query(query, (str(record_id),))
            if not results:
                return False

            # Định dạng dữ liệu tiêu đề và các dòng
            headers = ["Tên thiết bị", "Số lượng", "Tình trạng trả", "Ngày mượn", "Hẹn trả"]
            rows = []
            
            borrower_name = results[0]['full_name']
            
            for row in results:
                status = "Đã trả" if row['returned'] else "Chưa trả"
                rows.append([
                    row['device_name'],
                    row['quantity'],
                    status,
                    row['borrow_date'][:10] if row['borrow_date'] else '',
                    row['expected_return_date'][:10] if row['expected_return_date'] else ''
                ])

            # Chèn thêm thông tin metadata vào đầu báo cáo
            final_rows = [
                ["MÃ ĐƠN MƯỢN:", str(record_id)],
                ["NGƯỜI MƯỢN:", borrower_name],
                ["NGÀY XUẤT HÓA ĐƠN:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                [], # Dòng trống phân cách
                headers
            ] + rows

            filename = os.path.join(self.export_dir, self.create_report_filename(f"Invoice_{record_id}"))
            
            # Sử dụng luồng xuất dữ liệu nội bộ
            if openpyxl is None or filename.endswith('.csv'):
                self._export_to_csv(headers=[], rows=final_rows, filename=filename)
            else:
                self._export_to_excel_raw(final_rows, filename)
                
            return True
        except Exception as e:
            print(f"Lỗi truy xuất hóa đơn #{record_id}: {e}")
            return False

    def export_borrowers_list_to_csv(self):
        """Truy xuất danh sách tất cả các thiết bị đang được mượn và chưa trả."""
        query = """
            SELECT br.record_id, u.full_name, d.device_name, bd.quantity, 
                   br.borrow_date, br.expected_return_date, br.status
            FROM borrow_records br
            JOIN users u ON br.user_id = u.user_id
            JOIN borrow_details bd ON br.record_id = bd.record_id
            JOIN devices d ON bd.device_id = d.device_id
            WHERE bd.returned = 0 AND br.status IN ('Đã phê duyệt', 'Đang mượn', 'Quá hạn')
            ORDER BY br.expected_return_date ASC
        """
        try:
            results = self.db.execute_query(query)
            if not results:
                return False

            headers = ["Mã Đơn", "Người Mượn", "Tên Thiết Bị", "Số Lượng", "Ngày Mượn", "Ngày Hẹn Trả", "Trạng Thái"]
            rows = []
            for row in results:
                rows.append([
                    row['record_id'],
                    row['full_name'],
                    row['device_name'],
                    row['quantity'],
                    row['borrow_date'][:10] if row['borrow_date'] else '',
                    row['expected_return_date'][:10] if row['expected_return_date'] else '',
                    row['status']
                ])

            filename = os.path.join(self.export_dir, self.create_report_filename("Active_Borrowers_List"))
            self.export_to_excel(headers, rows, filename)
            return True
        except Exception as e:
            print(f"Lỗi xuất danh sách người mượn: {e}")
            return False

    # --- CÁC HÀM TIỆN ÍCH XỬ LÝ ĐỊNH DẠNG FILE ---

    def export_to_excel(self, headers, rows, filename):
        """Xuất mảng dữ liệu cơ bản thành file Excel với header bôi đậm."""
        if openpyxl is None:
            self._export_to_csv(headers, rows, filename)
            return f"Đã xuất dữ liệu sang CSV: {filename}"

        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Report"

            bold = Font(bold=True)
            for col_index, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col_index, value=header)
                cell.font = bold

            for row_index, row in enumerate(rows, start=2):
                for col_index, value in enumerate(row, start=1):
                    worksheet.cell(row=row_index, column=col_index, value=value)

            workbook.save(filename)
            return f"Báo cáo đã được lưu vào: {filename}"
        except Exception as e:
            print(f"Lỗi lưu file Excel: {e}")
            self._export_to_csv(headers, rows, filename)
            return f"Chuyển đổi dự phòng, đã lưu CSV: {filename}"

    def _export_to_excel_raw(self, matrix_rows, filename):
        """Xuất mảng dữ liệu thô (có chứa metadata) ra file Excel không theo định dạng chuẩn."""
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Invoice"

            for row_index, row in enumerate(matrix_rows, start=1):
                for col_index, value in enumerate(row, start=1):
                    worksheet.cell(row=row_index, column=col_index, value=value)

            workbook.save(filename)
        except Exception as e:
            print(f"Lỗi ghi file Excel raw: {e}")

    def _export_to_csv(self, headers, rows, filename):
        """Hàm dự phòng xuất ra CSV, hỗ trợ cả trường hợp có hoặc không có Header độc lập."""
        csv_filename = os.path.splitext(filename)[0] + ".csv"
        try:
            with open(csv_filename, mode="w", newline="", encoding="utf-8-sig") as file:
                writer = csv.writer(file)
                if headers:
                    writer.writerow(headers)
                writer.writerows(rows)
        except Exception as e:
            print(f"Lỗi phân bổ I/O khi lưu CSV: {e}")

    def create_report_filename(self, prefix="report"):
        """Sinh tên file chuẩn hóa."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = ".xlsx" if openpyxl else ".csv"
        return f"{prefix}_{timestamp}{extension}"